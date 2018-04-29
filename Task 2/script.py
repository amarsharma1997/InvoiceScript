from openpyxl import Workbook,load_workbook
import inflect
from shutil import copyfile

statedict={}
orders = {}
statecode = {}


class Lineitem:
    def __init__(self,list):
        self.listdetail={}
        self.listdetail['name']=list['Lineitem name']
        self.listdetail['HSN'] = '6101'
        self.listdetail['quantity']=list['Lineitem quantity']
        self.listdetail['rate'] = list['Lineitem price']
        self.listdetail['per'] = 'Nos'
        self.listdetail['amount'] = str(int(list['Lineitem quantity'])*int(list['Lineitem price']))



class orderDetails:
    def __init__(self,list):
        self.details={}
        self.details['name'] = list['Name']
        self.details['paymode'] = list['Payment Method']
        self.details['date'] = list['Created at'].date()
        self.details['billname'] = list['Billing Name']
        self.details['address'] = list['Shipping Address1'] + ',' + list['Shipping City']
        self.details['mobile'] = 'Mobile :' + list['Shipping Phone']
        self.details['state'] = statedict[list['Shipping Province']].upper()
        self.details['statecode'] = statecode[self.details['state'].lower()]
        self.details['GST'] = list['Taxes']
        self.details['total'] = list['Total']
        self.details['amount'] = "In words : "+inflect.engine().number_to_words(list['Total'])+" ONLY"
        self.details['list'] = []

    def addinlist(self,list):
        neworder = Lineitem(list)
        self.details['list'].append(neworder)

    def __str__(self):
        return self.details['name'] + ' ' + str(self.details['Total'])

    def storeorders(self,worksheet,side):
        if(side == 0):
            i=0
            for orderdetails in self.details['list']:
                worksheet['A' + str(i + 23) ].value = str(i+1)
                worksheet['B' + str(i + 23)].value = orderdetails.listdetail['name']
                worksheet['F' + str(i + 23)].value = orderdetails.listdetail['HSN']
                worksheet['G' + str(i + 23)].value = orderdetails.listdetail['quantity']
                worksheet['H' + str(i + 23)].value = orderdetails.listdetail['rate']
                worksheet['I' + str(i + 23)].value = orderdetails.listdetail['per']
                worksheet['J' + str(i + 23)].value = orderdetails.listdetail['amount']
                i=i+1
        else:
            i = 0
            for orderdetails in self.details['list']:
                worksheet['M' + str(i + 23)].value = str(i + 1)
                worksheet['N' + str(i + 23)].value = orderdetails.listdetail['name']
                worksheet['R' + str(i + 23)].value = orderdetails.listdetail['HSN']
                worksheet['S' + str(i + 23)].value = orderdetails.listdetail['quantity']
                worksheet['T' + str(i + 23)].value = orderdetails.listdetail['rate']
                worksheet['U' + str(i + 23)].value = orderdetails.listdetail['per']
                worksheet['V' + str(i + 23)].value = orderdetails.listdetail['amount']
                i = i + 1

    def store(self,list,worksheet,side):
        for key,value in list.items():
            worksheet[key].value = self.details[value]
        self.storeorders(worksheet,side)




def readAllStateAbb():
    wb = load_workbook('state.xlsx')
    ws = wb['Sheet1']
    for i in range(2,38):
        statedict[ws['B' + str(i)].value]=ws['A'+str(i)].value.lower()




def readAllStateCode():
    wb = load_workbook('GST.xlsx')
    ws = wb['Sheet1']
    for i in range(2, 38):
        statecode[ws['B' + str(i)].value.lower()] = ws['C' + str(i)].value




def readAllOrders():
    wb = load_workbook('orders.xlsx',data_only=True)
    ws = wb['orders']
    string = 'ABCDEFGHIJKLMNOPQ'
    fieldnames = []
    for s in string:
        fieldnames.append(ws[s + '1'].value)
    i=2
    while(True):
        if(ws['A'+str(i)].value == None):
            break
        name=ws['A'+str(i)].value
        list={}
        for j in range(len(string)):
            list[fieldnames[j]] = ws[string[j]+str(i)].value
        if(orders.get(name)==None):
            orders[name] = orderDetails(list)
        orders[name].addinlist(list)
        i = i+1




def createInvoices():
    total=0
    wb = load_workbook('sampleinvoice.xlsx')
    ws = wb['Sheet1']
    list = {}
    list[0]={}
    list[1]={}
    index=[]
    index.append("ABCDEFGHIJK")
    index.append("MNOPQRSTUVW")
    for row in index[0]:
        for col in range(1,38):
            if(ws[row + str(col)].value == None):
                 continue
            string = ws[row + str(col)].value.split(' ')
            if string[0] == 'xxx':
                list[0][row + str(col)] = string[1]

    for row in index[1]:
        for col in range(1,38):
            if(ws[row + str(col)].value == None):
                continue
            string = ws[row + str(col)].value.split(' ')
            if string[0] == 'xxx':
                list[1][row + str(col)] = string[1]
    wb.close()

    for i in range(0,len(orders),2):
        wb = Workbook()
        newname = 'invoiceoutput' + str(int((i + 2) / 2)) + '.xlsx'
        wb.save(newname)
        wb.close()
        copyfile('sampleinvoice.xlsx', newname)

    for name,details in orders.items():
        newname = 'invoiceoutput' + str(int((total + 2) / 2)) + '.xlsx'
        if(total%2 == 0):
            wb = load_workbook(newname)
            ws = wb.worksheets[0]
            details.store(list[0],ws,0)
            wb.save(newname)
        else:
            details.store(list[1], ws,1)
            wb.save(newname)
            wb.close()
        total = total+1




readAllStateAbb()
readAllStateCode()
readAllOrders()
createInvoices()